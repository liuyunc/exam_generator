# main.py

import asyncio
import json
import os
import re
import logging
from io import BytesIO
from typing import Callable, List, Optional
from pathlib import Path

from fastapi import FastAPI, UploadFile, Form, Request, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware import Middleware
from starlette.middleware.base import BaseHTTPMiddleware
from pydantic import BaseModel, Field, conint, constr
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from openai import (
    APIConnectionError,
    APIError,
    APITimeoutError,
    AuthenticationError,
    OpenAI,
)

from prompts import GA_SYSTEM_PROMPT, build_ga_user_prompt
from docx_utils import build_docx_from_ga, sort_ga_pairs_by_type
from json_validator import JSONValidator, validate_and_sanitize_ga_response

# ========= 日志配置 =========
logger = logging.getLogger(__name__)

# ========= 安全配置常量 =========
MIN_TIMEOUT = 10
MAX_TIMEOUT = 600  # 10分钟
MIN_RETRIES = 1
MAX_RETRIES = 5
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
MAX_JSON_SIZE = 5 * 1024 * 1024  # 5MB
MAX_QUESTIONS = 100
MIN_QUESTIONS = 1

ALLOWED_STATIC_DIR = Path("static").resolve()
ALLOWED_INDEX_FILE = ALLOWED_STATIC_DIR / "index.html"

# ========= 配置：GPUStack/DeepSeek =========


def _clean_env_value(raw: str, name: str) -> str:
    """去除环境变量两端空格，避免配置中出现看不见的前缀/后缀。"""

    if raw is None:
        return ""

    cleaned = raw.strip()
    if cleaned != raw:
        logger.warning(f"[config] 环境变量 {name} 含有首尾空格，已自动去除。")
    return cleaned


def _normalize_base_url(raw: str, name: str) -> str:
    """标准化 base_url：去除空格与末尾斜杠，兼容 DeepSeek 直连或 GPUStack 代理。"""

    cleaned = _clean_env_value(raw, name)
    if not cleaned:
        raise ValueError(f"环境变量 {name} 不能为空")
    
    if cleaned.endswith("/"):
        cleaned = cleaned[:-1]
    
    # 基础验证
    if not cleaned.startswith("http://") and not cleaned.startswith("https://"):
        raise ValueError(f"base_url 必须以 http:// 或 https:// 开头")
    
    return cleaned


def _get_first_env(names: list, *, required_name: str) -> str:
    """按优先级读取多个环境变量，便于兼容不同命名。强制要求设置。"""

    for key in names:
        raw = os.getenv(key)
        if raw:
            if key != required_name:
                logger.info(
                    f"[config] 检测到 {key}，已作为 {required_name} 使用（建议改为 {required_name} 以避免混淆）。"
                )
            return _clean_env_value(raw, key)

    raise ValueError(
        f"必须设置环境变量 {required_name}（或等价变量 {', '.join(names)}）"
    )


try:
    GPUSTACK_API_KEY = _get_first_env(
        ["GPUSTACK_API_KEY", "DEEPSEEK_API_KEY"],
        required_name="GPUSTACK_API_KEY",
    )
    GPUSTACK_BASE_URL = _normalize_base_url(
        _get_first_env(
            ["GPUSTACK_BASE_URL", "DEEPSEEK_BASE_URL"],
            required_name="GPUSTACK_BASE_URL",
        ),
        "GPUSTACK_BASE_URL",
    )
except ValueError as e:
    logger.error(f"[config] 初始化失败: {e}")
    raise

MODEL_NAME = _clean_env_value(
    os.getenv("DEEPSEEK_MODEL_NAME", "deepseek-r1"), "DEEPSEEK_MODEL_NAME"
)

# 添加超时和重试的范围限制
raw_timeout = float(
    _clean_env_value(os.getenv("GPUSTACK_TIMEOUT", "120"), "GPUSTACK_TIMEOUT") or "120"
)
GPUSTACK_TIMEOUT = max(MIN_TIMEOUT, min(raw_timeout, MAX_TIMEOUT))
if GPUSTACK_TIMEOUT != raw_timeout:
    logger.warning(f"[config] GPUSTACK_TIMEOUT 已调整到 [{MIN_TIMEOUT}, {MAX_TIMEOUT}] 范围")

raw_retries = int(
    _clean_env_value(os.getenv("GPUSTACK_MAX_RETRIES", "2"), "GPUSTACK_MAX_RETRIES") or "2"
)
GPUSTACK_MAX_RETRIES = max(MIN_RETRIES, min(raw_retries, MAX_RETRIES))
if GPUSTACK_MAX_RETRIES != raw_retries:
    logger.warning(f"[config] GPUSTACK_MAX_RETRIES 已调整到 [{MIN_RETRIES}, {MAX_RETRIES}] 范围")

logger.info(
    f"[config] base_url={GPUSTACK_BASE_URL}, model={MODEL_NAME}, "
    f"timeout={GPUSTACK_TIMEOUT}s, retries={GPUSTACK_MAX_RETRIES}"
)

client = OpenAI(
    api_key=GPUSTACK_API_KEY,
    base_url=GPUSTACK_BASE_URL,
)

# ========= FastAPI 应用 =========

app = FastAPI(title="JSON分片考试题生成器（DeepSeek + GA对）")

# 添加 CORS 中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=os.getenv("CORS_ORIGINS", "http://localhost:3000,http://localhost:8000").split(","),
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)


# 安全头中间件
class SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "1; mode=block"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
        return response


app.add_middleware(SecurityHeadersMiddleware)

# 静态文件目录：static/index.html
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/", response_class=HTMLResponse)
async def index():
    """返回前端页面（安全版本）"""
    try:
        if not ALLOWED_INDEX_FILE.exists():
            logger.error(f"index.html 文件不存在: {ALLOWED_INDEX_FILE}")
            return HTMLResponse("<h1>页面加载失败</h1>", status_code=500)
        
        with open(ALLOWED_INDEX_FILE, "r", encoding="utf-8") as f:
            html = f.read()
        return HTMLResponse(html)
    except Exception as e:
        logger.exception("加载 index.html 失败")
        return HTMLResponse("<h1>页面加载失败，请稍后重试</h1>", status_code=500)


# ========= 数据模型 =========

class GAPair(BaseModel):
    id: Optional[str] = None
    question_type: Optional[str] = ""
    options: Optional[List[str]] = Field(default_factory=list)
    question: str
    ga_answer: str
    score: Optional[float] = None
    tag: Optional[str] = ""
    difficulty: Optional[str] = ""
    source_excerpt: Optional[str] = ""
    source_locator: Optional[str] = ""
    comment: Optional[str] = ""


class ExportDocxRequest(BaseModel):
    title: constr(max_length=200)
    ga_pairs: List[GAPair] = Field(max_items=1000)


class ExportXlsxRequest(BaseModel):
    ga_pairs: List[GAPair] = Field(max_items=1000)


class GARequest(BaseModel):
    """纯 API 调用版本（非网页上传）"""
    chunks: List[dict] = Field(min_items=1, max_items=1000)
    chunk_indices: List[int] = Field(max_items=100)
    num_questions: conint(ge=MIN_QUESTIONS, le=MAX_QUESTIONS) = 20
    system_prompt: Optional[constr(max_length=5000)] = None


# ========= 工具函数 =========

def extract_chunk_items(chunks: list, indices: list):
    """
    根据索引只抽取需要的分片，返回：
    [
      {
        "index": i,
        "title": "xxx",
        "text": "这一分片的正文"
      },
      ...
    ]
    """
    items = []
    for i in indices:
        if i < 0 or i >= len(chunks):
            continue
        item = chunks[i]
        text = (
            item.get("content")
            or item.get("text")
            or item.get("chunk")
            or ""
        )
        title = item.get("name") or item.get("fileName") or f"chunk-{i}"
        items.append({
            "index": i,
            "title": title,
            "text": str(text).strip()
        })
    return items


def extract_json_block_from_content(content: str) -> dict:
    """
    从大模型返回的 content 文本中，稳健地抽取出一个 JSON 对象（带安全检查）。
    优先寻找以 {"ga_pairs" 开头的 JSON；如果没有，就从第一个 { 开始做括号匹配。
    """
    if not content:
        raise ValueError("模型返回内容为空，无法解析 JSON")

    if len(content) > 10 * MAX_JSON_SIZE:
        raise ValueError(f"模型返回内容过大（超过 {10 * MAX_JSON_SIZE} 字节）")

    # 1) 优先找 {"ga_pairs"
    start = content.find('{"ga_pairs"')
    if start == -1:
        # 退而求其次：找第一个 {
        start = content.find("{")
    if start == -1:
        raise ValueError("未在模型返回中找到 '{'，可能没有输出 JSON")

    in_str = False
    escape = False
    depth = 0
    end = None

    for i in range(start, len(content)):
        ch = content[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = i + 1
                    break

    if end is None:
        end = len(content)

    json_str = content[start:end].strip()
    
    if not json_str:
        raise ValueError("提取到的 JSON 字符串为空")

    if len(json_str) > MAX_JSON_SIZE:
        raise ValueError(f"JSON 块大小超过 {MAX_JSON_SIZE} 字节限制")

    try:
        data = json.loads(json_str)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 解析失败: {str(e)}")

    # 验证结构
    if not isinstance(data, dict):
        raise ValueError("JSON 必须是一个对象")
    
    if "ga_pairs" not in data:
        raise ValueError("JSON 缺少必需的 'ga_pairs' 字段")
    
    if not isinstance(data["ga_pairs"], list):
        raise ValueError("'ga_pairs' 字段必须是列表")

    return data


def call_deepseek_ga_single_chunk(
    text_for_model: str,
    num_questions: int,
    system_prompt: Optional[str] = None,
    log_fn: Callable[[str], None] = print,
):
    """针对单个分片调用 DeepSeek 生成 GA 对（带更稳健的 JSON 解析与验证）"""
    sys_prompt = system_prompt.strip() if system_prompt else GA_SYSTEM_PROMPT
    user_prompt = build_ga_user_prompt(text_for_model, num_questions)

    resp = None
    last_error = None
    for attempt in range(1, GPUSTACK_MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": sys_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0.3,
                timeout=GPUSTACK_TIMEOUT,
            )
            break
        except (APITimeoutError, APIConnectionError) as e:
            last_error = "API 调用超时或连接异常"
            logger.warning(
                f"[DeepSeek] 第 {attempt}/{GPUSTACK_MAX_RETRIES} 次调用超时/连接异常；"
                f"超时设置 {GPUSTACK_TIMEOUT}s"
            )
            if attempt == GPUSTACK_MAX_RETRIES:
                return [], last_error
            time.sleep(min(2 * attempt, 6))
        except AuthenticationError as e:
            last_error = "认证失败，请检查 API 密钥"
            logger.error(f"[DeepSeek] 认证失败")
            return [], last_error
        except APIError as e:
            last_error = "API 服务异常，请稍后重试"
            logger.error(f"[DeepSeek] 服务器返回错误: {type(e).__name__}")
            return [], last_error
        except Exception as e:
            last_error = "API 调用失败，请稍后重试"
            logger.exception("API 调用异常")
            return [], last_error

    if resp is None:
        return [], last_error or "调用失败，未返回响应"

    content = resp.choices[0].message.content or ""

    try:
        # 先尝试直接当 JSON 解析
        data = json.loads(content)
    except json.JSONDecodeError:
        # 如果失败，则用括号匹配从 content 中提取 JSON 块
        try:
            data = extract_json_block_from_content(content)
        except Exception as e:
            error_msg = f"模型返回内容无法解析为 JSON"
            logger.warning(f"JSON 解析失败: {str(e)}")
            logger.debug(f"模型原始返回（前 500 字符）: {content[:500]}")
            return [], error_msg

    # ========= 新增：JSON 验证与幻觉检测 =========
    ga_pairs, validation_result = validate_and_sanitize_ga_response(data, strict_mode=False)
    
    # 记录验证信息
    log_fn(JSONValidator.log_validation_report(validation_result, len(data.get("ga_pairs", []))))
    
    if not validation_result.is_valid:
        error_msg = f"JSON 验证失败：{'; '.join(validation_result.errors[:3])}"
        logger.error(f"JSON 验证失败: {validation_result.errors}")
        return [], error_msg
    
    logger.info(f"JSON 验证通过，信任度评分: {validation_result.score:.2%}")
    return ga_pairs, None


@app.get("/api/deepseek-health")
async def deepseek_health():
    """快速检测 DeepSeek/GPUStack 连接与鉴权是否正常。"""

    start_ts = time.time()
    try:
        resp = client.models.list()
        elapsed_ms = int((time.time() - start_ts) * 1000)
        model_ids = []
        try:
            model_ids = [m.id for m in getattr(resp, "data", [])][:3]
        except Exception:
            model_ids = []

        return {
            "ok": True,
            "message": f"连接成功，耗时 {elapsed_ms} ms",
            "models": model_ids,
        }
    except AuthenticationError:
        return {"ok": False, "message": "认证失败，请检查 API 密钥"}
    except (APIConnectionError, APITimeoutError):
        return {"ok": False, "message": "连接失败或超时"}
    except APIError:
        return {"ok": False, "message": "服务异常"}
    except Exception as e:
        logger.exception("健康检查异常")
        return {"ok": False, "message": "健康检查失败"}


@app.get("/api/system-prompt")
async def get_system_prompt():
    """返回后端默认的 System Prompt，便于前端展示与编辑。"""

    return {"system_prompt": GA_SYSTEM_PROMPT}


def call_deepseek_ga_for_chunks(
    chunk_items: list,
    total_questions: int,
    system_prompt: Optional[str] = None,
    log_fn: Callable[[str], None] = print,
):
    """
    按分片分别调用 DeepSeek，再汇总 GA 对：
    - total_questions：总题量
    - 各分片按数量平均分配
    """
    if not chunk_items or total_questions <= 0:
        return [], ["未提供分片或题目数量小于等于 0"]

    n_chunks = len(chunk_items)
    base = total_questions // n_chunks
    rem = total_questions % n_chunks

    all_pairs = []
    errors: List[str] = []
    for idx, item in enumerate(chunk_items):
        n_q = base + (1 if idx < rem else 0)
        if n_q <= 0:
            continue

        log_fn(f"正在处理分片{item['index']}（{item['title']}），预计生成{n_q}道题目...")
        
        header = f"[分片{item['index']}：{item['title']}]\n"
        text_for_model = header + item["text"]

        ga_pairs, error_msg = call_deepseek_ga_single_chunk(
            text_for_model=text_for_model,
            num_questions=n_q,
            system_prompt=system_prompt,
            log_fn=log_fn,
        )

        if error_msg:
            errors.append(
                f"分片{item['index']}（{item['title']}）调用 DeepSeek 失败：{error_msg}"
            )
            log_fn(errors[-1])
            continue

        log_fn(f"分片{item['index']}处理完成，实际生成{len(ga_pairs)}道题目")

        # 给每个 GA 对附加分片定位（兜底）
        for p in ga_pairs:
            locator = (p.get("source_locator") or "").strip()
            extra = f"（自动定位：分片{item['index']}，{item['title']}）"
            if locator:
                locator = locator + "；" + extra
            else:
                locator = extra
            p["source_locator"] = locator
        all_pairs.extend(ga_pairs)
    
    log_fn(f"所有分片处理完成，共生成{len(all_pairs)}道题目")

    return all_pairs, errors


def _normalize_question_type(question_type: Optional[str]) -> str:
    return (question_type or "").strip().lower()


def _is_single_choice(question_type: Optional[str]) -> bool:
    normalized = _normalize_question_type(question_type)
    raw = question_type or ""
    return "single" in normalized or "单选" in raw or "single_choice" in normalized


def _is_multiple_choice(question_type: Optional[str]) -> bool:
    normalized = _normalize_question_type(question_type)
    raw = question_type or ""
    return "multiple" in normalized or "多选" in raw or "multiple_choice" in normalized


def _normalize_options(options: Optional[List[str] | str]) -> List[str]:
    if options is None:
        return []
    if isinstance(options, str):
        raw_options = options.splitlines()
    else:
        raw_options = options
    cleaned = []
    for opt in raw_options:
        if opt is None:
            continue
        text = str(opt).strip()
        if not text:
            continue
        text = re.sub(r"^[A-Ha-h][\.\:：、\)]\s*", "", text)
        cleaned.append(text)
    return cleaned


def _build_analysis_text(pair: GAPair) -> str:
    parts = []
    if pair.difficulty:
        parts.append(f"【难度】{_sanitize_math_markdown(pair.difficulty)}")
    if pair.source_locator:
        parts.append(f"【来源定位】{_sanitize_math_markdown(pair.source_locator)}")
    if pair.source_excerpt:
        parts.append(f"【原文摘录】{_sanitize_math_markdown(pair.source_excerpt)}")
    if pair.comment:
        parts.append(f"【命题说明】{_sanitize_math_markdown(pair.comment)}")
    return "  ".join(parts)


def _sanitize_math_markdown(text: str) -> str:
    """将 Markdown/LaTeX 形式的公式转为可读文本，避免 Excel 中保留 $ 符号。"""

    if not text:
        return ""

    def _clean_math_content(content: str) -> str:
        content = re.sub(r"\\mathrm\{([^}]+)\}", r"\1", content)
        content = re.sub(r"\\operatorname\{([^}]+)\}", r"\1", content)
        content = re.sub(r"_\{([^}]+)\}", r"_\1", content)
        content = re.sub(r"\^\{([^}]+)\}", r"^\1", content)
        content = re.sub(r"\\text\{([^}]+)\}", r"\1", content)
        content = re.sub(r"\\(left|right|bigl|bigr|Bigl|Bigr|biggl|biggr|Biggl|Biggr)", "", content)
        content = re.sub(r"\\([a-zA-Z]+)", r"\1", content)
        return content.replace("{", "").replace("}", "")

    inline_patterns = [
        r"\$\$(.+?)\$\$",
        r"\$(.+?)\$",
        r"\\\((.+?)\\\)",
        r"\\\[(.+?)\\\]",
    ]
    for pattern in inline_patterns:
        text = re.sub(pattern, lambda m: _clean_math_content(m.group(1)), text)
    return _clean_math_content(text)


def build_xlsx_from_ga(ga_pairs: List[GAPair]) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "序号",
        "题干",
        "选项 A",
        "选项 B",
        "选项 C",
        "选项 D",
        "选项 E",
        "选项 F",
        "选项 G",
        "选项 H",
        "解析",
        "分数",
        "答案",
        "标签",
    ]

    header_fill = PatternFill("solid", fgColor="FFF200")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def add_sheet(title: str, rows: List[GAPair]):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        for index, pair in enumerate(rows, start=1):
            options = _normalize_options(pair.options)
            option_cells = options[:8] + [""] * max(0, 8 - len(options))
            analysis = _build_analysis_text(pair)
            score = "" if pair.score is None else pair.score
            question = _sanitize_math_markdown(pair.question)
            answer = _sanitize_math_markdown(pair.ga_answer)
            tag = _sanitize_math_markdown(pair.tag or "")
            row = [
                index,
                question,
                *[_sanitize_math_markdown(opt) for opt in option_cells],
                analysis,
                score,
                answer,
                tag,
            ]
            ws.append(row)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = cell_alignment
                cell.border = border

        widths = [6, 42, 22, 22, 22, 22, 22, 22, 22, 22, 48, 8, 10, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

        ws.freeze_panes = "A2"

    single_pairs = [p for p in ga_pairs if _is_single_choice(p.question_type)]
    multiple_pairs = [p for p in ga_pairs if _is_multiple_choice(p.question_type)]
    add_sheet("单选题", single_pairs)
    add_sheet("多选题", multiple_pairs)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ========= API：网页调用 =========

@app.post("/api/generate-ga-from-file")
async def generate_ga_from_file(
    file: UploadFile,
    chunk_indices: str = Form(
        "",
        description="分片索引，如：0,1,2；留空则自动使用全部分片",
    ),
    num_questions: int = Form(20),
    system_prompt: str = Form("", description="自定义 system prompt，可留空使用默认"),
):
    """
    网页表单接口（支持流式返回日志）：
    - 上传 JSON 分片文件
    - 指定要使用的分片索引（逗号分隔）
    - 指定题目总数量
    - 可选：自定义提示词
    """

    # 验证文件类型
    if file.content_type not in ["application/json", "application/octet-stream", "text/plain"]:
        logger.warning(f"不允许的文件类型: {file.content_type}")
        raise HTTPException(
            status_code=400, 
            detail="仅支持 JSON 文件"
        )

    # 用于前端实时展示：将日志/结果放入队列，StreamingResponse 边产生边下发
    queue: asyncio.Queue[str | None] = asyncio.Queue()
    logs: List[str] = []
    loop = asyncio.get_running_loop()

    def enqueue(payload: dict):
        """把事件写入队列，前端按行解析（线程安全）。"""

        loop.call_soon_threadsafe(
            queue.put_nowait, json.dumps(payload, ensure_ascii=False)
        )

    def log_and_collect(msg: str):
        logs.append(str(msg))
        logger.info(msg)
        enqueue({"type": "log", "message": str(msg)})

    raw = await file.read()

    # 检查文件大小
    if len(raw) > MAX_FILE_SIZE:
        error_msg = f"文件超过 {MAX_FILE_SIZE} 字节限制"
        logger.warning(f"文件过大: {len(raw)} 字节")
        raise HTTPException(status_code=413, detail=error_msg)

    if len(raw) == 0:
        raise HTTPException(status_code=400, detail="文件为空")

    async def producer():
        try:
            def generate():
                log_and_collect("开始处理文件上传…")
                
                try:
                    chunks = json.loads(raw)
                except json.JSONDecodeError as e:
                    error_msg = f"文件不是有效的 JSON: {str(e)}"
                    log_and_collect(error_msg)
                    enqueue({"type": "error", "message": error_msg, "logs": logs})
                    return

                log_and_collect("文件解析完成")

                # 支持：顶层是 {'chunks': [...]} 或直接是 list
                if isinstance(chunks, dict) and "chunks" in chunks:
                    chunks_list = chunks["chunks"]
                elif isinstance(chunks, list):
                    chunks_list = chunks
                else:
                    error_msg = "文件格式错误：需要 JSON 数组或对象包含 'chunks' 字段"
                    log_and_collect(error_msg)
                    enqueue({"type": "error", "message": error_msg, "logs": logs})
                    return

                # 验证题目数量
                if not MIN_QUESTIONS <= num_questions <= MAX_QUESTIONS:
                    error_msg = f"题目数量必须在 {MIN_QUESTIONS}-{MAX_QUESTIONS} 之间"
                    log_and_collect(error_msg)
                    enqueue({"type": "error", "message": error_msg, "logs": logs})
                    return

                # 解析索引
                indices = []
                for part in chunk_indices.split(","):
                    part = part.strip()
                    if not part:
                        continue
                    try:
                        idx = int(part)
                        indices.append(idx)
                    except ValueError:
                        continue

                if not indices:
                    indices = list(range(len(chunks_list)))
                    log_and_collect(
                        f"未显式指定分片索引，自动使用全部 {len(chunks_list)} 个分片: {indices}"
                    )
                else:
                    log_and_collect(f"解析到 {len(indices)} 个分片索引: {indices}")
                
                chunk_items = extract_chunk_items(chunks_list, indices)
                log_and_collect(f"提取到 {len(chunk_items)} 个有效分片")

                ga_pairs, errors = call_deepseek_ga_for_chunks(
                    chunk_items,
                    total_questions=num_questions,
                    system_prompt=system_prompt if system_prompt.strip() else None,
                    log_fn=log_and_collect,
                )

                log_and_collect(
                    f"生成完成，共生成 {len(ga_pairs)} 道题目；错误 {len(errors)} 条"
                )

                enqueue(
                    {
                        "type": "result",
                        "ga_pairs": ga_pairs,
                        "errors": errors,
                        "logs": logs,
                    }
                )

            # 在线程中执行耗时/阻塞的同步 DeepSeek 调用，避免卡住事件循环导致前端无法即时收到日志
            await asyncio.to_thread(generate)
        except Exception as e:
            error_msg = "处理文件时发生错误，请稍后重试"
            logger.exception("文件处理异常")
            log_and_collect(error_msg)
            enqueue({"type": "error", "message": error_msg, "logs": logs})
        finally:
            # 结束标记
            await queue.put(None)

    asyncio.create_task(producer())

    async def event_stream():
        while True:
            item = await queue.get()
            if item is None:
                break
            yield item + "\n"

    return StreamingResponse(event_stream(), media_type="application/x-ndjson")


@app.post("/export-docx")
async def export_docx(req: ExportDocxRequest):
    """接收前端编辑好的 GA 对，生成 DOCX 下载"""
    try:
        ga_pairs_dicts = [p.dict() for p in req.ga_pairs]
        sorted_ga_pairs = sort_ga_pairs_by_type(ga_pairs_dicts)
        doc = build_docx_from_ga(sorted_ga_pairs, title=req.title)

        buffer = BytesIO()
        doc.save(buffer)
        buffer.seek(0)

        filename = "exam_ga_pairs.docx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers=headers,
        )
    except Exception as e:
        logger.exception("导出 DOCX 异常")
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")


@app.post("/export-xlsx")
async def export_xlsx(req: ExportXlsxRequest):
    """接收前端编辑好的 GA 对，生成 XLSX 下载"""
    try:
        output = build_xlsx_from_ga(req.ga_pairs)
        filename = "exam_ga_pairs.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        logger.exception("导出 XLSX 异常")
        raise HTTPException(status_code=500, detail="导出失败，请稍后重试")


# ========= API：纯后端调用版（可选） =========

@app.post("/api/generate-ga")
async def api_generate_ga(req: GARequest):
    """
    纯 JSON API（不走上传文件），方便后续和 EasyDataset pipeline 联动
    """
    try:
        chunk_items = extract_chunk_items(req.chunks, req.chunk_indices)
        ga_pairs, errors = call_deepseek_ga_for_chunks(
            chunk_items,
            total_questions=req.num_questions,
            system_prompt=req.system_prompt,
        )
        return {"ga_pairs": ga_pairs, "errors": errors}
    except Exception as e:
        logger.exception("API 生成异常")
        raise HTTPException(status_code=500, detail="生成失败，请稍后重试")
