from io import BytesIO

from openpyxl import load_workbook

from docx_utils import build_docx_from_ga, sort_ga_pairs_by_type
from main import GAPair, build_xlsx_from_ga


def sample_pairs():
    return [
        {
            "id": "q2",
            "question_type": "multiple_choice",
            "options": ["A. One", "B. Two", "C. Three", "D. Four"],
            "question": "Choose values",
            "ga_answer": "AC",
            "difficulty": "medium",
            "source_excerpt": "source text",
            "source_locator": "chunk-1",
            "comment": "checks recall",
        },
        {
            "id": "q1",
            "question_type": "single_choice",
            "options": ["A. Yes", "B. No", "C. Maybe", "D. Unknown"],
            "question": "Choose one",
            "ga_answer": "A",
            "difficulty": "easy",
            "source_excerpt": "source text",
            "source_locator": "chunk-0",
            "comment": "checks basics",
        },
    ]


def test_sort_ga_pairs_by_type_keeps_choice_order():
    sorted_pairs = sort_ga_pairs_by_type(sample_pairs())

    assert [pair["id"] for pair in sorted_pairs] == ["q1", "q2"]


def test_build_docx_from_ga_creates_document():
    doc = build_docx_from_ga(sample_pairs(), title="Assessment")
    output = BytesIO()

    doc.save(output)

    assert output.tell() > 0
    assert doc.paragraphs[0].text == "Assessment"


def test_build_xlsx_from_ga_creates_choice_sheets():
    pairs = [GAPair(**pair) for pair in sample_pairs()]

    output = build_xlsx_from_ga(pairs)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["单选题", "多选题"]
    assert workbook["单选题"].max_row == 2
    assert workbook["多选题"].max_row == 2
